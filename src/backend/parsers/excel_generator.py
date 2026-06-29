"""
excel_generator.py
Generates a populated module data sheet Excel file from extracted JSON data.
Matches the layout of "module data sheet .xlsx".

Usage:
    from excel_generator import generate_module_excel

    data = {
        "manufacturer": "SunPower",
        "module_model": "SPR-MAX3-400",
        "bifacial_coefficient": 0.7,
        "noct": 45,
        "series_fuse_rating": 30,
        "operating_temperature_range": "-40°C to 85°C",
        "dimensions_mm": "2382 x 1134 x 33",
        "weight_kg": 42.5,
        "cells_count": 132,
        "cell_type": "Monocrystalline PERC",
        "front_glass": "3.2mm Tempered",
        "back_glass": "2.0mm Tempered",
        "output_cable": "4mm² 300mm",
        "connector": "MC4",
        "junction_box": "IP68",
        "temperature_coefficients": {"isc_alpha": 0.046, "voc_beta": -0.25, "pm_gamma": -0.30},
        "load_rating": {"wind": 2400, "snow": 5400},
        "degradation": {"first_year": 2.0, "yearly": 0.55, "year_30": None},
        "warranty": {"product": 12, "performance": 30},
        "variants": [
            {"pmax": 400, "pstc": 400, "voc": 49.5, "vmp": 41.8, "isc": 10.2, "imp": 9.57, "efficiency": 20.4}
        ]
    }

    output_path = generate_module_excel(data, "output_module_sheet.xlsx")
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from typing import Optional
import os


# ── Colour palette ──────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "2E75B6"   # mid-blue
SUBHDR_FG   = "FFFFFF"
ALT_ROW_BG  = "DEEAF1"   # light-blue alternate rows
UNIT_BG     = "F2F2F2"
BORDER_CLR  = "9DC3E6"


def _border(clr=BORDER_CLR):
    s = Side(style="thin", color=clr)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_font(bold=True, color=HEADER_FG, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Main function ────────────────────────────────────────────────────────────

def generate_module_excel(data: dict, output_path: str = "module_data_sheet.xlsx") -> str:
    """
    Build and save a formatted module data sheet Excel workbook.

    Parameters
    ----------
    data : dict
        Extracted module data matching the extractor.py JSON schema.
    output_path : str
        Destination path for the .xlsx file.

    Returns
    -------
    str
        Absolute path to the created file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Module Data Sheet"

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = {1: 8, 2: 38, 3: 22, 4: 22, 5: 22, 6: 22, 7: 22, 8: 22, 9: 12}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # ── Title row ────────────────────────────────────────────────────────────
    manufacturer = data.get("manufacturer") or "—"
    module_model = data.get("module_model") or "—"
    title = f"PV Module Data Sheet — {manufacturer} | {module_model}"

    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value = title
    tc.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=14)
    tc.fill = _fill(HEADER_BG)
    tc.alignment = _center()
    tc.border = _border()

    # ── Column header row ────────────────────────────────────────────────────
    variants = data.get("variants") or []
    num_variants = max(len(variants), 1)

    col_headers = ["Sr. No", "Technical Parameters", "Unit"]
    for i in range(1, num_variants + 1):
        pmax = variants[i - 1].get("pmax") if i <= len(variants) else "—"
        col_headers.append(f"Variant {i}" + (f"\n({pmax} Wp)" if pmax else ""))

    for ci, hdr in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font = _hdr_font()
        cell.fill = _fill(SUBHDR_BG)
        cell.alignment = _center(wrap=True)
        cell.border = _border()

    # ── Helper: write a data row ─────────────────────────────────────────────
    def write_row(row_num: int, sr: Optional[int], label: str, unit: str,
                  value, variant_key: Optional[str] = None, alt: bool = False):
        bg = ALT_ROW_BG if alt else "FFFFFF"
        ws.row_dimensions[row_num].height = 18

        # Sr. No
        c = ws.cell(row=row_num, column=1, value=sr if sr is not None else "")
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()

        # Label
        c = ws.cell(row=row_num, column=2, value=label)
        c.font = Font(name="Calibri", size=10)
        c.fill = _fill(bg)
        c.alignment = _left()
        c.border = _border()

        # Unit
        c = ws.cell(row=row_num, column=3, value=unit)
        c.font = Font(name="Calibri", size=10, italic=True)
        c.fill = _fill(UNIT_BG)
        c.alignment = _center()
        c.border = _border()

        # Values — if variant_key provided, pull from each variant; else repeat value
        for vi in range(1, num_variants + 1):
            col_idx = 3 + vi
            if variant_key and vi <= len(variants):
                val = variants[vi - 1].get(variant_key)
            else:
                val = value
            c = ws.cell(row=row_num, column=col_idx, value=val if val is not None else "—")
            c.font = Font(name="Calibri", size=10)
            c.fill = _fill(bg)
            c.alignment = _center()
            c.border = _border()

    # ── Helper: section divider ───────────────────────────────────────────────
    def write_section(row_num: int, label: str):
        ws.row_dimensions[row_num].height = 16
        ws.merge_cells(f"A{row_num}:I{row_num}")
        c = ws.cell(row=row_num, column=1, value=f"  {label}")
        c.font = Font(name="Calibri", bold=True, color=SUBHDR_FG, size=10)
        c.fill = _fill(SUBHDR_BG)
        c.alignment = _left(wrap=False)
        c.border = _border()

    # ── Data rows ────────────────────────────────────────────────────────────
    tc_raw = data.get("temperature_coefficients") or {}
    load_raw = data.get("load_rating") or {}
    deg_raw = data.get("degradation") or {}
    war_raw = data.get("warranty") or {}

    r = 3  # current row pointer

    # — Electrical Parameters (per-variant) ——
    write_section(r, "⚡  Electrical Parameters (STC: 1000 W/m², 25°C, AM1.5)"); r += 1

    write_row(r, 1,  "Module Rated Power (Pmax)",   "Wp",    None, "pmax",       alt=False); r += 1
    write_row(r, 2,  "Pstc (Front Side)",            "Wp",    None, "pstc",       alt=True);  r += 1
    write_row(r, 3,  "Voc (Front Side)",              "V",     None, "voc",        alt=False); r += 1
    write_row(r, 4,  "Vmp (Front Side)",              "V",     None, "vmp",        alt=True);  r += 1
    write_row(r, 5,  "Isc (Front Side)",              "A",     None, "isc",        alt=False); r += 1
    write_row(r, 6,  "Imp (Front Side)",              "A",     None, "imp",        alt=True);  r += 1
    write_row(r, 7,  "Module Efficiency",             "%",     None, "efficiency", alt=False); r += 1

    # — Thermal Parameters ——
    write_section(r, "🌡  Thermal Parameters"); r += 1

    write_row(r, 8,  "Temp. Coefficient of Current (Isc), α",  "%/°C", tc_raw.get("isc_alpha"), alt=False); r += 1
    write_row(r, 9,  "Temp. Coefficient of Voltage (Voc), β",  "%/°C", tc_raw.get("voc_beta"),  alt=True);  r += 1
    write_row(r, 10, "Temp. Coefficient of Power (Pm), γ",     "%/°C", tc_raw.get("pm_gamma"),  alt=False); r += 1
    write_row(r, 11, "NOCT",                                    "°C",   data.get("noct"),        alt=True);  r += 1

    # — Mechanical Parameters ——
    write_section(r, "🔩  Mechanical & Physical Parameters"); r += 1

    write_row(r, 12, "Operating Temperature Range",    "°C",  data.get("operating_temperature_range"), alt=False); r += 1
    write_row(r, 13, "Series Fuse Max Rating",         "A",   data.get("series_fuse_rating"),           alt=True);  r += 1
    write_row(r, 14, "Dimensions (L × W × H)",         "mm",  data.get("dimensions_mm"),                alt=False); r += 1
    write_row(r, 15, "Weight",                         "Kg",   data.get("weight_kg"),                   alt=True);  r += 1
    write_row(r, 16, "Solar Cells per Module",         "Units",data.get("cells_count"),                 alt=False); r += 1
    write_row(r, 17, "Solar Cell Type",                "—",    data.get("cell_type"),                   alt=True);  r += 1

    # — Construction ——
    write_section(r, "🏗  Construction & Components"); r += 1

    write_row(r, 18, "Front Glass",       "—", data.get("front_glass"),  alt=False); r += 1
    write_row(r, 19, "Back Glass",        "—", data.get("back_glass"),   alt=True);  r += 1
    write_row(r, 20, "Output Cable",      "—", data.get("output_cable"), alt=False); r += 1
    write_row(r, 21, "Connector",         "—", data.get("connector"),    alt=True);  r += 1
    write_row(r, 22, "Junction Box",      "—", data.get("junction_box"), alt=False); r += 1

    # — Load Rating ——
    write_section(r, "💨  Load Rating"); r += 1

    write_row(r, 23, "Wind Load",  "Pa", load_raw.get("wind"), alt=False); r += 1
    write_row(r, 24, "Snow Load",  "Pa", load_raw.get("snow"), alt=True);  r += 1

    # — Degradation ——
    write_section(r, "📉  Degradation & Warranty"); r += 1

    write_row(r, 25, "First Year Degradation",    "%",       deg_raw.get("first_year"), alt=False); r += 1
    write_row(r, 26, "Annual Degradation",         "%/year",  deg_raw.get("yearly"),     alt=True);  r += 1
    write_row(r, 27, "30-Year Degradation",        "%",       deg_raw.get("year_30"),    alt=False); r += 1
    write_row(r, 28, "Product Warranty",           "Years",   war_raw.get("product"),    alt=True);  r += 1
    write_row(r, 29, "Performance Warranty",       "Years",   war_raw.get("performance"),alt=False); r += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1,
                value="Generated by Forge Extractor Pipeline  |  PVInsight Inc.  |  Data sourced from manufacturer datasheet PDF")
    c.font = Font(name="Calibri", italic=True, color="888888", size=9)
    c.alignment = _center()

    # ── Freeze panes & print setup ────────────────────────────────────────────
    ws.freeze_panes = "C3"
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    abs_path = os.path.abspath(output_path)
    wb.save(abs_path)
    print(f"✅ Excel saved → {abs_path}")
    return abs_path


# ── CLI demo ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sample_data = {
        "manufacturer": "SunPower",
        "module_model": "SPR-MAX3-400",
        "bifacial_coefficient": 0.7,
        "noct": 45,
        "series_fuse_rating": 30,
        "operating_temperature_range": "-40°C to 85°C",
        "dimensions_mm": "2382 x 1134 x 33",
        "weight_kg": 42.5,
        "cells_count": 132,
        "cell_type": "Monocrystalline PERC",
        "front_glass": "3.2mm Tempered",
        "back_glass": "2.0mm Tempered",
        "output_cable": "4mm² / 300mm",
        "connector": "MC4",
        "junction_box": "IP68 Rated",
        "temperature_coefficients": {
            "isc_alpha": 0.046,
            "voc_beta": -0.25,
            "pm_gamma": -0.30
        },
        "load_rating": {"wind": 2400, "snow": 5400},
        "degradation": {"first_year": 2.0, "yearly": 0.55, "year_30": 83.1},
        "warranty": {"product": 12, "performance": 30},
        "variants": [
            {"pmax": 390, "pstc": 390, "voc": 48.8, "vmp": 41.1, "isc": 10.1, "imp": 9.49, "efficiency": 19.8},
            {"pmax": 395, "pstc": 395, "voc": 49.1, "vmp": 41.5, "isc": 10.1, "imp": 9.52, "efficiency": 20.1},
            {"pmax": 400, "pstc": 400, "voc": 49.5, "vmp": 41.8, "isc": 10.2, "imp": 9.57, "efficiency": 20.4},
        ]
    }

    generate_module_excel(sample_data, "module_data_sheet_output.xlsx")
